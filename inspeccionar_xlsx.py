"""
Utilidad de inspección: muestra hojas y primeras filas de un xlsx,
para que puedas decidir qué hoja/columnas mapear en app/mappings.py

Uso:
    python -m app.inspeccionar_xlsx "corpus_v1/01_factores_generales/Factor de Emision.xlsx"
    python -m app.inspeccionar_xlsx "ruta.xlsx" --hoja "CO2 EFs" --filas 20
"""
import argparse
import openpyxl


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--hoja", default=None, help="Nombre de hoja específica a inspeccionar")
    parser.add_argument("--filas", type=int, default=15, help="Cantidad de filas a mostrar")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.path, data_only=True, read_only=True)

    if not args.hoja:
        print(f"Hojas disponibles en {args.path}:")
        for nombre in wb.sheetnames:
            print(f"  - {nombre}")
        print("\nVolvé a correr con --hoja \"<nombre>\" para ver el contenido de una hoja.")
        return

    ws = wb[args.hoja]
    print(f"Primeras {args.filas} filas de la hoja '{args.hoja}':\n")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=args.filas, values_only=True), start=1):
        # Recortamos filas muy largas para que se lea bien en consola
        fila_recortada = row[:10]
        print(f"{i:>3}: {fila_recortada}")


if __name__ == "__main__":
    main()

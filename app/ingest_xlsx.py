"""
Ingesta factores de emisión estructurados desde los xlsx del corpus,
usando EXCLUSIVAMENTE lo declarado en app/mappings.py (ver ese archivo
para entender por qué no es un parser automático).

Uso:
    python -m app.ingest_xlsx
"""
from pathlib import Path
import openpyxl
from tqdm import tqdm

from app.config import CORPUS_PATH, CORPUS_VERSION
from app.db import get_connection, insert_factor
from app.mappings import MAPEOS_HOJAS_VALIDAS


def _buscar_archivo(nombre_archivo: str) -> Path:
    coincidencias = list(Path(CORPUS_PATH).rglob(nombre_archivo))
    if not coincidencias:
        raise FileNotFoundError(f"No se encontró '{nombre_archivo}' dentro de {CORPUS_PATH}")
    return coincidencias[0]


def ingestar_mapeo(conn, mapeo: dict):
    path = _buscar_archivo(mapeo["archivo"])
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if mapeo["hoja"] not in wb.sheetnames:
        print(f"  ! Hoja '{mapeo['hoja']}' no existe en {mapeo['archivo']} - se omite")
        return 0

    ws = wb[mapeo["hoja"]]
    cols = mapeo["columnas"]
    insertados = 0

    for i, row in enumerate(
        ws.iter_rows(min_row=mapeo["fila_inicio"], values_only=True), start=mapeo["fila_inicio"]
    ):
        nombre_factor = row[cols["nombre_factor"]] if cols["nombre_factor"] is not None else None
        valor = row[cols["valor"]] if cols["valor"] is not None else None

        # Salteamos filas vacías o sin valor numérico (evita basura en la tabla)
        if nombre_factor is None or valor is None:
            continue
        try:
            valor = float(valor)
        except (TypeError, ValueError):
            continue

        unidad = row[cols["unidad"]] if cols.get("unidad") is not None else None

        insert_factor(
            conn,
            archivo_origen=mapeo["archivo"],
            hoja=mapeo["hoja"],
            fila=i,
            nombre_factor=str(nombre_factor).strip(),
            valor=valor,
            unidad=str(unidad).strip() if unidad else None,
            gas=mapeo.get("gas_fijo"),
            fuente=mapeo.get("fuente_fija"),
            anio=mapeo.get("anio_fijo"),
            alcance=mapeo["alcance"],
            categoria=mapeo["categoria"],
            version_corpus=CORPUS_VERSION,
        )
        insertados += 1

    return insertados


def main():
    if not MAPEOS_HOJAS_VALIDAS:
        print("app/mappings.py está vacío. Agregá al menos un mapeo antes de correr esto.")
        return

    conn = get_connection()
    total = 0
    try:
        for mapeo in tqdm(MAPEOS_HOJAS_VALIDAS, desc="Ingestando factores xlsx"):
            total += ingestar_mapeo(conn, mapeo)
    finally:
        conn.close()

    print(f"Listo. {total} factores insertados en `factores_emision` (versión {CORPUS_VERSION}).")


if __name__ == "__main__":
    main()
